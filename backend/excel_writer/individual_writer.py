# excel_writer/individual_writer.py

from openpyxl.styles import Alignment, Font
from ieee519 import individual_current_limit


def write_individual_sheet(wb, individual_list, transformer_list=None):
    """
    Creates the 'Individual Harmonics' sheet and fills H1–H13 spectrum values.

    Red highlighting (IEEE 519):
      - Each individual harmonic cell I1–I13 is coloured red if its value
        exceeds the IEEE 519 individual current harmonic limit for that
        harmonic order and the Isc/IL ratio of the matching transformer row.

    Feeder Name cells are merged vertically for consecutive rows
    that share the same feeder name.
    """

    ws = wb.create_sheet("Individual Harmonics")

    red = Font(color="FF0000")

    # ── Build lookup: (feeder, type) → transformer row ───────────────────────
    t_map = {}
    if transformer_list:
        for t in transformer_list:
            # individual_list uses "APFC ON/OFF" for the type key
            key = (t.get("Feeder_Name"), t.get("Type"))
            t_map[key] = t

    # ── Header ────────────────────────────────────────────────────────────────
    ws.append(
        ["S.No", "Feeder", "Type"] +
        [f"I{i}" for i in range(1, 14)]
    )

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, r in enumerate(individual_list, 1):
        ws.append(
            [i, r.get("Feeder Name"), r.get("APFC ON/OFF")] +
            [r.get(f"I{x}", 0) for x in range(1, 14)]
        )

        data_row = ws.max_row

        # Match transformer row by feeder + type
        key = (r.get("Feeder Name"), r.get("APFC ON/OFF"))
        t   = t_map.get(key) or {}

        try:
            isc_il = float(t.get("Isc/IL") or 0)
        except (TypeError, ValueError):
            isc_il = 0

        # ── Red: each I1–I13 cell if > individual limit ───────────────────────
        # I1 is column D (index 4), I2 = col E (5), … I13 = col P (16)
        for order in range(1, 14):
            col_idx = 3 + order   # col D=4 for order 1
            try:
                val   = float(r.get(f"I{order}") or 0)
                limit = individual_current_limit(isc_il, order)
                if val > limit:
                    ws.cell(row=data_row, column=col_idx).font = red
            except (TypeError, ValueError):
                pass

    # ── Merge Feeder (col B) for consecutive same-feeder data rows ────────────
    # Data starts at Excel row 2 (row 1 is header)
    DATA_START = 2
    n = len(individual_list)

    i = 0
    while i < n:
        feeder = individual_list[i].get("Feeder Name")

        j = i + 1
        while j < n and individual_list[j].get("Feeder Name") == feeder:
            j += 1

        if j - i > 1:
            row_start = DATA_START + i
            row_end   = DATA_START + j - 1
            ws.merge_cells(f"B{row_start}:B{row_end}")
            ws[f"B{row_start}"].alignment = Alignment(
                vertical="center", wrap_text=True
            )

        i = j

    return ws
