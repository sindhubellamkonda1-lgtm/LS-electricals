# excel_writer/harmonic_writer.py

from openpyxl.styles import Alignment, Font
from ieee519 import tdd_limit, voltage_thd_limit


def write_harmonic_sheet(wb, harmonic_list, transformer_list=None):
    """
    Creates the 'Harmonic Analysis' sheet and fills harmonic summary data.

    Red highlighting (IEEE 519):
      - Current THD (col K): red if Max_A_THD > TDD limit (from Isc/IL)
      - Voltage THD (col L): red if Max_U_THD > voltage THD limit (from lv_kv)

    Feeder Name cells are merged vertically for consecutive rows
    that share the same feeder name.
    """

    ws = wb.create_sheet("Harmonic Analysis")

    red = Font(color="FF0000")

    # ── Build a quick lookup: (feeder, type) → transformer row ───────────────
    t_map = {}
    if transformer_list:
        for t in transformer_list:
            key = (t.get("Feeder_Name"), t.get("Type"))
            t_map[key] = t

    # ── Header ────────────────────────────────────────────────────────────────
    ws.append([
        "S.No",
        "Feeder",
        "Type",
        "Current RMS",
        "Voltage RMS",
        "KW",
        "KVAR",
        "KVA",
        "PF",
        "DPF",
        "Current THD",
        "Voltage THD",
    ])

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, r in enumerate(harmonic_list, 1):
        ws.append([
            i,
            r.get("Feeder_Name"),
            r.get("Type"),
            r.get("Max_A"),
            r.get("Max_U_RMS"),
            r.get("KW"),
            r.get("KVAR"),
            r.get("KVA"),
            r.get("PF"),
            r.get("DPF"),
            r.get("Max_A_THD"),
            r.get("Max_U_THD"),
        ])

        data_row = ws.max_row

        # Look up matching transformer row to get Isc/IL and lv_kv
        key = (r.get("Feeder_Name"), r.get("Type"))
        t   = t_map.get(key) or {}

        # ── Red: Current THD (col K = 11) > TDD limit ─────────────────────────
        try:
            isc_il    = float(t.get("Isc/IL") or 0)
            max_a_thd = float(r.get("Max_A_THD") or 0)
            if max_a_thd > tdd_limit(isc_il):
                ws.cell(row=data_row, column=11).font = red
        except (TypeError, ValueError):
            pass

        # ── Red: Voltage THD (col L = 12) > voltage limit ─────────────────────
        try:
            lv_kv     = float(t.get("lv_kv") or 0.433)
            max_u_thd = float(r.get("Max_U_THD") or 0)
            if max_u_thd > voltage_thd_limit(lv_kv):
                ws.cell(row=data_row, column=12).font = red
        except (TypeError, ValueError):
            pass

    # ── Merge Feeder (col B) for consecutive same-feeder data rows ────────────
    # Data starts at Excel row 2 (row 1 is header)
    DATA_START = 2
    n = len(harmonic_list)

    i = 0
    while i < n:
        feeder = harmonic_list[i].get("Feeder_Name")

        j = i + 1
        while j < n and harmonic_list[j].get("Feeder_Name") == feeder:
            j += 1

        if j - i > 1:
            row_start = DATA_START + i
            row_end   = DATA_START + j - 1
            ws.merge_cells(f"B{row_start}:B{row_end}")
            ws[f"B{row_start}"].alignment = Alignment(
                vertical="center", wrap_text=True
            )

        i = j

    return ws
