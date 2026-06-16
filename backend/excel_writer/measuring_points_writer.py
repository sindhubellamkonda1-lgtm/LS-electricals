# excel_writer/measuring_points_writer.py

from openpyxl.styles import Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


def write_feeder_sheet(wb, data):
    """
    Creates the 'Feeder Report' sheet
    and fills basic feeder + capacitor information.
    The dynamic values in the APFC Remarks (kVAr amount and reactor status)
    are rendered in red to highlight them.
    """

    # Create or use active sheet
    ws = wb.active
    ws.title = "Measuring Points"

    # Extract JSON info
    site = data["site_measurements"][0]
    feeder = site["feeder_name"]
    cap = site["capacitor_bank_details"]

    rating_kvar   = cap.get("rating_kvar", "?")
    reactor_value = "Yes" if cap.get("reactors_present") else "Nil"

    # Header row
    ws.append(["S.No", "Feeder", "Remarks"])

    # Row 1
    ws.append([
        1,
        feeder,
        "Recording with Capacitor ON and OFF condition"
    ])

    # Row 2 — append with blank Remarks placeholder, then set rich text below
    ws.append([2, feeder, ""])

    # Build rich-text Remarks: plain text + red values
    red = InlineFont(color="FF0000")
    ws["C3"] = CellRichText(
        "APFC Banks \u2013 ",
        TextBlock(red, f"{rating_kvar} kVAr"),
        ", Reactors \u2013 ",
        TextBlock(red, reactor_value)
    )

    # Merge Feeder cells (B2:B3) when both rows share the same feeder name
    ws.merge_cells("B2:B3")
    ws["B2"].alignment = Alignment(vertical="center", wrap_text=True)

    return ws
