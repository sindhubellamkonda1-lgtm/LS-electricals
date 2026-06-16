# excel_writer/transformer_writer.py

from openpyxl.styles import Alignment, Font


# ── Header definition ──────────────────────────────────────────────────────────
# Each tuple: (column header text, unit string or None)
# If unit is None → rows 1 & 2 are merged, text rotated vertical & bold.
# If unit has a value → row 1 = header (bold), row 2 = unit (bold), no merge.

COLUMNS = [
    ("S.No",                                          None),
    ("Feeder Name",                                   None),
    ("APFC ON/OFF",                                   None),
    ("Transformer Capacity",                          "kVA"),
    ("Voltage Rating",                                "KV"),
    ("Full load current (Transformer rating)",        "Amps"),
    ("Impedance %(Z)",                                "%"),
    ("Vector Group",                                  None),
    ("Short circuit current (Isc)",                   "Amps"),
    ("Transformer full load amps (IL Measured)",      "Amps"),
    ("Isc/I L",                                       None),
    ("TDD -total demand distortion (IEEE standard)",  None),
    ("Ithd %",                                        None),
]


def write_transformer_sheet(wb, transformer_list):
    """
    Creates the 'Transformer Limits' sheet.

    Layout:
      Row 1  — column header names
      Row 2  — units sub-row  (kVA, KV, Amps, %, …)
      Row 3+ — data rows

    Header formatting:
      - Columns whose unit cell is empty → rows 1 & 2 merged,
        text rotated 90° (vertical), bold, centred.
      - Columns with a unit → row 1 header bold + centred,
        row 2 unit bold + centred, no merge.

    Feeder Name cells are merged vertically for consecutive data rows
    that share the same feeder name.
    """

    ws = wb.create_sheet("Transformer Limits")

    bold            = Font(bold=True)
    vertical_bold   = Alignment(textRotation=90, horizontal="center", vertical="center")
    centred         = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Row 1: header names ───────────────────────────────────────────────────
    ws.append([col for col, _ in COLUMNS])

    # ── Row 2: units ──────────────────────────────────────────────────────────
    ws.append([unit for _, unit in COLUMNS])

    # ── Apply header formatting ───────────────────────────────────────────────
    for col_idx, (header, unit) in enumerate(COLUMNS, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter

        if unit is None:
            # Merge rows 1 & 2, rotate text vertical, bold
            ws.merge_cells(f"{col_letter}1:{col_letter}2")
            cell = ws[f"{col_letter}1"]
            cell.font      = bold
            cell.alignment = vertical_bold
        else:
            # Header row — bold, centred
            h_cell = ws[f"{col_letter}1"]
            h_cell.font      = bold
            h_cell.alignment = centred
            # Unit row — bold, centred
            u_cell = ws[f"{col_letter}2"]
            u_cell.font      = bold
            u_cell.alignment = centred

    red = Font(color="FF0000")

    # ── Rows 3+: data ─────────────────────────────────────────────────────────
    for i, r in enumerate(transformer_list, 1):
        ws.append([
            i,
            r.get("Feeder_Name"),
            r.get("Type"),
            r.get("Transformer_Capacity_KVA"),
            r.get("voltage_rate_kva"),
            r.get("full_load_amps"),
            r.get("impedance_percent"),
            r.get("vector_group"),
            r.get("short_cirtuit_amps"),
            r.get("Transformer_A"),
            r.get("Isc/IL"),
            r.get("TDD"),
            r.get("Ithd"),
        ])

        # ── Red highlight: Ithd% > TDD limit (col M = index 13) ───────────────
        data_row = ws.max_row
        try:
            ithd = float(r.get("Ithd") or 0)
            tdd  = float(r.get("TDD")  or 0)
            if ithd > tdd:
                ws.cell(row=data_row, column=13).font = red
        except (TypeError, ValueError):
            pass

    # ── Merge Feeder Name (col B) for consecutive same-feeder data rows ───────
    # Data starts at Excel row 3 (rows 1-2 are header + units)
    DATA_START = 3
    n = len(transformer_list)

    i = 0
    while i < n:
        feeder = transformer_list[i].get("Feeder_Name")

        j = i + 1
        while j < n and transformer_list[j].get("Feeder_Name") == feeder:
            j += 1

        if j - i > 1:
            row_start = DATA_START + i
            row_end   = DATA_START + j - 1
            ws.merge_cells(f"B{row_start}:B{row_end}")
            ws[f"B{row_start}"].alignment = Alignment(
                vertical="center", wrap_text=True
            )

        i = j

    return ws
