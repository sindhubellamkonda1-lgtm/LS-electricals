import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from datetime import datetime
from utils_numeric import num
from excel_loader import load_sheet_fast
from datetime_utils import normalize_excel_datetime
from harmonic_engine import Harmonic_Analysis, extract_harmonic_max
from transformer_engine import Trasformer_Limits_Analysis
from window_detector import build_max_power_df





# ==========================================================
# POWER WINDOW DETECTION
# ==========================================================

# ==========================================================
# HARMONIC SUMMARY ⭐ FULL DATA
# ==========================================================


# ==========================================================
# GENERATE XLS ⭐ FULL REPORT + GRAPHS
# ==========================================================
def Generate_XLS(data,max_df,transformer_list,harmonic_list,individual_list,img_dir="."):

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    import os
    from datetime import datetime

    name=f"Transformer_Report_{datetime.now().strftime('%H%M%S')}.xlsx"
    wb=Workbook()

    # ---------------------------------------------------------
    # TAB 1 : FEEDER REPORT
    # ---------------------------------------------------------
    ws=wb.active
    ws.title="Feeder Report"

    site=data["site_measurements"][0]
    feeder=site["feeder_name"]
    cap=site["capacitor_bank_details"]

    ws.append(["S.No","Feeder","Remarks"])

    ws.append([
        1,
        feeder,
        "Recording with Capacitor ON and OFF condition"
    ])

    ws.append([
        2,
        feeder,
        f"APFC Banks – {cap.get('rating_kvar','?')} kVAr, "
        f"Reactors – {'Yes' if cap.get('reactors_present') else 'Nil'}"
    ])

    # ---------------------------------------------------------
    # TAB 2 : TRANSFORMER LIMITS (FULL DATA RESTORED)
    # ---------------------------------------------------------
    ws1=wb.create_sheet("Transformer Limits")

    ws1.append([
        "S.No","Feeder Name","APFC ON/OFF",
        "Transformer Capacity","Voltage Rating",
        "Full load current",
        "Impedance %",
        "Vector Group",
        "Short circuit current",
        "Measured Current",
        "Isc/IL",
        "TDD",
        "Ithd"
    ])

    for i,r in enumerate(transformer_list,1):

        ws1.append([
            i,
            r.get("Feeder_Name"),
            r.get("Type"),
            r.get("Transformer_Capacity_KVA"),
            r.get("voltage_rate_kva"),
            r.get("full_load_amps"),
            r.get("impedance_percent"),
            r.get("vector_group"),
            r.get("short_cirtuit_amps"),
            r.get("Transformer_A"),
            r.get("Isc/IL"),
            r.get("TDD"),
            r.get("Ithd")
        ])

    # ---------------------------------------------------------
    # TAB 3 : HARMONIC ANALYSIS (FULL DATA RESTORED)
    # ---------------------------------------------------------
    ws2=wb.create_sheet("Harmonic Analysis")

    ws2.append([
        "S.No","Feeder","Type","Date","Time",
        "Current RMS",
        "Voltage RMS",
        "KW",
        "KVAR",
        "KVA",
        "PF",
        "DPF",
        "Current THD",
        "Voltage THD"
    ])

    for i,r in enumerate(harmonic_list,1):

        ws2.append([
            i,
            r.get("Feeder_Name"),
            r.get("Type"),
            r.get("Date"),
            r.get("Time"),
            r.get("Max_A"),
            r.get("Max_U_RMS"),
            r.get("KW"),
            r.get("KVAR"),
            r.get("KVA"),
            r.get("PF"),
            r.get("DPF"),
            r.get("Max_A_THD"),
            r.get("Max_U_THD")
        ])

    # ---------------------------------------------------------
    # TAB 4 : INDIVIDUAL HARMONICS
    # ---------------------------------------------------------
    ws3=wb.create_sheet("Individual Harmonics")

    ws3.append(["S.No","Feeder","Type"]+[f"I{i}" for i in range(1,14)])

    for i,r in enumerate(individual_list,1):

        ws3.append(
            [i,r.get("Feeder Name"),r.get("APFC ON/OFF")]
            +[r.get(f"I{x}",0) for x in range(1,14)]
        )

    # ---------------------------------------------------------
    # TAB 5 : GRAPHS
    # ---------------------------------------------------------
    ws_img=wb.create_sheet("Graphs")

    row=3

    for title,file in [
        ("Load Profile","load_curve.png"),
        ("THD","thd.png"),
        ("Loading","loading.png"),
        ("Harmonic Spectrum","harmonic_spectrum.png"),
        ("Power Triangle","triangle.png")
    ]:

        path=os.path.join(img_dir,file)

        if os.path.exists(path):

            ws_img[f"A{row}"]=title
            row+=1

            img=Image(path)
            img.width=900
            img.height=420

            ws_img.add_image(img,f"A{row}")
            row+=23

    wb.save(name)
    return name
